import requests
import openpyxl

def query_total_availability(url, headers, mpn):
    total_availability_query = '''
    query {
        supSearchMpn(q: "%s", limit: 1) {
            results {
                description
                part {
                    totalAvail
                    mpn
                }
            }
        }
    }
    ''' % mpn

    response = requests.post(url, headers=headers, json={"query": total_availability_query})
    return response.json()

def query_pricing_by_volume(url, headers, mpn):
    pricing_by_volume_query = '''
    query {
        supSearchMpn(q: "%s", limit: 1) {
            hits
            results {
                part {
                    totalAvail
                    sellers {
                        company {
                            name
                        }
                        offers {
                            prices {
                                quantity
                                price
                            }
                        }
                    }
                }
            }
        }
    }
    ''' % mpn

    response = requests.post(url, headers=headers, json={"query": pricing_by_volume_query})
    return response.json()

def process_mpn_data(url, headers, mpn_list, workbook):
    for mpn in mpn_list:
        sheet = create_sheet_if_not_exists(workbook, mpn)  # Create or retrieve the sheet

        data_total_availability = query_total_availability(url, headers, mpn)
        data_pricing_by_volume = query_pricing_by_volume(url, headers, mpn)

        # Print the response for debugging
        print("Total Availability Response:", data_total_availability)
        print("Pricing by Volume Response:", data_pricing_by_volume)

        # Check if 'data' key exists in the response
        if 'data' in data_total_availability:
            for result in data_total_availability['data']['supSearchMpn']['results']:
                description = result['description']
                mpn_value = result['part']['mpn']
                total_avail = result['part']['totalAvail']
                sheet.append([description, mpn_value, total_avail])
        else:
            print("Error: 'data' key not found in Total Availability response")

        # Write data to the sheet for pricing by volume
        if 'data' in data_pricing_by_volume:
            for result in data_pricing_by_volume['data']['supSearchMpn']['results']:
                for seller in result['part']['sellers']:
                    company_name = seller['company']['name']

                    for offer in seller['offers']:
                        # Updated this part to handle 'prices' as a list
                        for price_info in offer['prices']:
                            quantity = price_info['quantity']
                            price = price_info['price']

                            # Write data to the sheet
                            sheet.append(["", "", "", company_name, quantity, price])
        else:
            print("Error: 'data' key not found in Pricing by Volume response")

    # Save the workbook outside the loop
    workbook.save('exists.xlsx')

    print(data_total_availability)
    print(data_pricing_by_volume)

# Additional code to handle the case where the sheet already exists
def create_sheet_if_not_exists(workbook, sheet_name):
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(["Description", "MPN", "Total Availability", "Company Name", "Quantity", "Price"])

    return sheet

