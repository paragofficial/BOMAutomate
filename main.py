# main.py
# this is the final working code you can show it to sir 

import openpyxl
from mpn import process_mpn_data
from excel_function import process_excel_data_and_save


def main():
    url = "https://api.nexar.com/graphql"
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer eyJhbGciOiJSUzI1NiIsImtpZCI6IjA5NzI5QTkyRDU0RDlERjIyRDQzMENBMjNDNkI4QjJFIiwidHlwIjoiYXQrand0In0.eyJuYmYiOjE3MDcwNzI1MjMsImV4cCI6MTcwNzE1ODkyMywiaXNzIjoiaHR0cHM6Ly9pZGVudGl0eS5uZXhhci5jb20iLCJjbGllbnRfaWQiOiI4ODM5NDk0Yy02NTUwLTQwN2EtODQwYi01ZjI0M2I2OTRiMGIiLCJzdWIiOiJFN0M0MUJEMi0yQjQwLTQ4OTktOTRDMC05M0ExMDlGRjAzRDEiLCJhdXRoX3RpbWUiOjE3MDcwNzI0OTIsImlkcCI6ImxvY2FsIiwicHJpdmF0ZV9jbGFpbXNfaWQiOiIxNzdjNThjMC05MmY0LTQzZDAtYjc3Zi1jOTc3MGRmMTk1ZDciLCJwcml2YXRlX2NsYWltc19zZWNyZXQiOiJiV2k2RERvMjFLVEgyLzNCTHQ4WitzTGRYQmRabW1xZDNLM053U2NESDRFPSIsImp0aSI6IkI4Q0NGMUNDOTdFMDg5QjNBQTc0NjMxOURDMzk4RkQ1Iiwic2lkIjoiQTE5RUE0MjE4MzVGMUUyMkE3MjU3Q0Q4NkZCMTk2MUQiLCJpYXQiOjE3MDcwNzI1MjMsInNjb3BlIjpbIm9wZW5pZCIsInVzZXIuYWNjZXNzIiwicHJvZmlsZSIsImVtYWlsIiwidXNlci5kZXRhaWxzIiwic3VwcGx5LmRvbWFpbiIsImRlc2lnbi5kb21haW4iXSwiYW1yIjpbInB3ZCJdfQ.D2BnAHWX53TDVRQVuhEugdKqLIFtjanQm5iy3iXUzbnBCkg75z8hkB4j4QsjGpH-OHt1s-hu-eEtfLFJZM0rA4u-lls9JuCRncCVDHON6YmOwYe1D2NqPnN8uc25Sc7HkF98ofEqt4mm_MgeEzKE5dYkZCb7rjrwUxsgGWMLHm6MbFv9FfMBrOy6njIpw8qdADViC_JE49j25qiRVfulIPr8aeTVIP2Dj7Tr_Fr9jjkDB1fw_mSqSzLxIWOF8UXdWF2Koom028c1lTh8D4Cp60MXu8qGfpfdTSyOjaOCKLc3COvsV-v4pVFQy25EwOyPEZ5smnEeR6InG2Bt58NmeQ'
    }

    excel_file_path = 'RFQ_file.xlsx'
    sheet_name = 'J34A-BOMs'

    process_excel_data_and_save(excel_file_path, sheet_name, url, headers)

    # mpn_list = ["acs770", "atmega328p"]

    mpn_array = read_mpn_from_excel(file_path='exists.xlsx', sheet_name='Sheet1', column_name='MPN')
    print(mpn_array)

    workbook = openpyxl.load_workbook('exists.xlsx') 

    process_mpn_data(url, headers, mpn_array, workbook)



def read_mpn_from_excel(file_path='exists.xlsx', sheet_name='Sheet1', column_name='MPN'):
    mpn_values = []

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Find the column index for the specified column name
        column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            if col[0].value == column_name:
                column_index = col[0].column
                break

        if column_index is not None:
            # Iterate through the rows and extract MPN values
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
                mpn_values.append(row[0].value)

    except Exception as e:
        print(f"Error reading Excel file: {e}")

    return mpn_values


if __name__ == "__main__":
    main()



